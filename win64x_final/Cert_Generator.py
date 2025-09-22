# final/Cert_Generator.py
import openpyxl
import os
from datetime import datetime
import pandas as pd

def get_device_details(model_number):
    """
    Placeholder function to retrieve device details.
    """
    details = {
        "1350-00856": {"PART #": "1350-00856", "FITTING": "8fVCR", "CONNECTOR": "15 PIN"},
        "E28D-31351": {"PART #": "E28D-31351", "FITTING": "4fVCR", "CONNECTOR": "9 PIN"},
    }
    return details.get(model_number, {"PART #": "N/A", "FITTING": "N/A", "CONNECTOR": "N/A"})


def generate_certificate(dut_info, calibration_data, tech_id, log_queue):
    """
    Generates and saves a calibration certification Excel file.
    Returns the path to the generated file on success, otherwise None.
    """
    item_number = dut_info.get('item_number')
    wip = dut_info.get('wip', 'Unknown WIP')
    log_queue.put(f"Attempting to generate certificate for WIP {wip} using item number {item_number}.")

    if not item_number:
        log_queue.put(f"ERROR for WIP {wip}: No item number provided. Cannot generate certificate.")
        return None

    template_path = os.path.join('templates', f"{item_number}.xlsx")
    
    log_queue.put(f"Searching for template file at: {template_path}")

    if not os.path.exists(template_path):
        log_queue.put(f"ERROR for WIP {wip}: Template file not found at '{template_path}'.")
        return None

    log_queue.put(f"Template found for WIP {wip}. Proceeding with certificate generation.")
    
    output_dir = "Cal-Certs"
    os.makedirs(output_dir, exist_ok=True)

    try:
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        device_details = get_device_details(dut_info['model'])

        cell_map = {
            'G4': dut_info.get('customer', ''),
            'G5': device_details.get('PART #'),
            'G6': dut_info.get('wip', ''),
            'G7': datetime.now().strftime('%Y-%m-%d'),
            'G8': dut_info.get('model', ''),
            'G9': dut_info.get('serial', ''),
            'G10': f"{dut_info.get('fs')} TORR",
            'G11': device_details.get('FITTING'),
            'G12': device_details.get('CONNECTOR'),
            'G13': dut_info.get('cal_std_id', ''),
            'G14': "Vertical", 
            'G15': tech_id,
            'G16': dut_info.get('service_type', ''),
        }

        for cell_ref, value in cell_map.items():
            if value is not None:
                ws[cell_ref] = value

        dut_col_name = f'Device_{dut_info["channel"]+1}_Pressure_Torr'
        dut_cal_data = calibration_data[['Setpoint_Torr', 'Standard_Pressure_Torr', dut_col_name]].dropna()

        std_voltage = (dut_cal_data['Standard_Pressure_Torr'] / dut_info['fs']) * 10
        dut_voltage = (dut_cal_data[dut_col_name] / dut_info['fs']) * 10

        start_row = 21
        for i in range(len(dut_cal_data)):
            row = start_row + i
            ws[f'A{row}'] = 10 - i 
            ws[f'C{row}'] = std_voltage.iloc[i]
            ws[f'G{row}'] = dut_voltage.iloc[i]
            
        output_filename = f"{dut_info['wip']}.xlsx"
        output_path = os.path.join(output_dir, output_filename)
        wb.save(output_path)
        
        log_queue.put(f"✅ Successfully generated certificate for WIP {dut_info['wip']}. Saved to '{output_path}'")
        return output_path
        
    except Exception as e:
        log_queue.put(f"ERROR generating certificate for WIP {dut_info['wip']}: {e}")
        return None